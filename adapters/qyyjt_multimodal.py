#!/usr/bin/env python3
"""
adapters/qyyjt_multimodal.py — v1.1.0 企业预警通多模态文档处理

华尔街驻铁岭办事处 · 企业预警通图片/PDF/扫描件识别

企业预警通返回的数据格式多样:
  - 债券公告 → PDF (招股书/年报/评级报告)
  - 裁判文书 → PDF/扫描件
  - 财务报表 → HTML表格 / PDF / 图片截图
  - 企业证照 → 扫描件图片
  - 城投专题 → Excel / 图片图表

处理策略:
  1. 内容类型检测 (URL后缀 / HTTP Content-Type / 魔数)
  2. PDF文本提取 (PyPDF2 + pdfplumber)
  3. 图片OCR (Tesseract中文 + PaddleOCR兜底)
  4. 表格结构化提取 (PDF table → JSON)
  5. 营业执照/财报/裁判文书结构化字段提取
"""

from __future__ import annotations

import base64
import io
import json
import os
import re
import subprocess
import tempfile
import threading
from dataclasses import dataclass, field
from datetime import datetime
from enum import Enum
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple
from urllib.parse import urlparse

import requests


# ═══════════════════════════════════════════════════════════════
# 数据结构
# ═══════════════════════════════════════════════════════════════

class ContentType(Enum):
    """文件内容类型"""
    PDF = "pdf"
    IMAGE = "image"           # PNG/JPG/TIFF/BMP
    HTML = "html"             # HTML页面/表格
    EXCEL = "excel"           # XLS/XLSX/CSV
    TEXT = "text"             # 纯文本
    UNKNOWN = "unknown"


@dataclass
class ExtractedPage:
    page_number: int
    text: str = ""
    tables: List[List[List[str]]] = field(default_factory=list)
    confidence: float = 0.0
    is_scanned: bool = False


@dataclass
class ProcessedDocument:
    """多模态处理结果"""
    source_url: str
    content_type: ContentType
    total_pages: int = 0
    pages: List[ExtractedPage] = field(default_factory=list)
    full_text: str = ""
    structured_data: Dict[str, Any] = field(default_factory=dict)
    error: str = ""
    processing_time: float = 0.0


@dataclass
class EnterpriseLicense:
    """营业执照结构化数据"""
    uscc: str = ""              # 统一社会信用代码
    company_name: str = ""
    legal_person: str = ""
    registered_capital: str = ""
    established_date: str = ""
    business_scope: str = ""
    address: str = ""
    raw_text: str = ""


@dataclass
class CourtDocument:
    """裁判文书结构化数据"""
    case_number: str = ""       # 案号
    court: str = ""             # 审理法院
    case_type: str = ""         # 案件类型
    plaintiff: str = ""         # 原告
    defendant: str = ""         # 被告
    filing_date: str = ""       # 立案日期
    judgment_date: str = ""     # 判决日期
    judgment_amount: str = ""   # 涉案金额
    verdict: str = ""           # 判决结果
    raw_text: str = ""


@dataclass
class FinancialIndicators:
    """财务指标结构化数据"""
    total_assets: str = ""
    total_liabilities: str = ""
    revenue: str = ""
    net_profit: str = ""
    debt_ratio: str = ""
    current_ratio: str = ""
    roe: str = ""
    operating_cashflow: str = ""
    year: str = ""
    raw_text: str = ""


@dataclass
class PenaltyDecision:
    """行政处罚决定书结构化数据"""
    penalty_number: str = ""    # 处罚文号
    agency: str = ""            # 处罚机关
    target: str = ""            # 被处罚人
    violation: str = ""         # 违法事实
    penalty_type: str = ""      # 处罚种类
    penalty_amount: str = ""    # 罚款金额
    decision_date: str = ""     # 处罚日期
    legal_basis: str = ""       # 法律依据
    raw_text: str = ""


@dataclass
class AuditOpinion:
    """审计报告结构化数据"""
    auditor: str = ""           # 审计机构
    audit_type: str = ""        # 审计意见类型
    audit_year: str = ""        # 审计年度
    key_matters: str = ""       # 关键审计事项
    going_concern: str = ""     # 持续经营
    report_date: str = ""       # 报告日期
    raw_text: str = ""


@dataclass
class BondProspectus:
    """债券募集说明书结构化数据"""
    bond_name: str = ""         # 债券名称
    issuer: str = ""            # 发行人
    bond_type: str = ""         # 债券品种
    issue_amount: str = ""      # 发行规模
    maturity: str = ""          # 期限
    coupon_rate: str = ""       # 票面利率
    credit_rating: str = ""     # 信用评级
    lead_underwriter: str = ""  # 主承销商
    use_of_proceeds: str = ""   # 资金用途
    guarantor: str = ""         # 担保人
    raw_text: str = ""


@dataclass
class ExecutionRuling:
    """执行裁定书结构化数据"""
    case_number: str = ""       # 执行案号
    court: str = ""             # 执行法院
    applicant: str = ""         # 申请执行人
    respondent: str = ""        # 被执行人
    execution_amount: str = ""  # 执行标的
    filing_date: str = ""       # 立案日期
    ruling_date: str = ""       # 裁定日期
    execution_status: str = ""  # 执行情况
    property_found: str = ""    # 财产发现
    raw_text: str = ""


@dataclass
class DishonestyNotice:
    """失信被执行人决定书结构化数据"""
    case_number: str = ""       # 案号
    court: str = ""             # 执行法院
    dishonest_person: str = ""  # 失信被执行人
    id_number: str = ""         # 身份证/统一信用代码
    violation: str = ""         # 失信情形
    publish_date: str = ""      # 发布日期
    legal_basis: str = ""       # 法律依据
    raw_text: str = ""


@dataclass
class CreditRating:
    """信用评级报告结构化数据"""
    rated_entity: str = ""      # 评级对象
    rating_agency: str = ""     # 评级机构
    subject_rating: str = ""    # 主体信用等级
    bond_rating: str = ""       # 债项信用等级
    outlook: str = ""           # 评级展望
    rating_date: str = ""       # 评级日期
    key_factors: str = ""       # 评级关注
    upgrade_downgrade: str = "" # 评级调整
    raw_text: str = ""


# ═══════════════════════════════════════════════════════════════
# 主处理器
# ═══════════════════════════════════════════════════════════════

class QYJTMultimodalProcessor:
    """
    企业预警通多模态文档处理器

    用法:
        proc = QYJTMultimodalProcessor()

        # 处理URL
        doc = proc.process_url("https://www.qyyjt.cn/bond/pdf/xxx.pdf")

        # 处理字节流
        doc = proc.process_bytes(pdf_bytes, "年报.pdf")

        # 从API响应中智能提取
        doc = proc.extract_from_response(api_response)

        # 提取结构化信息
        if doc.content_type == ContentType.PDF:
            license_data = proc.extract_license(doc)
            print(license_data.company_name)
    """

    def __init__(self, ocr_lang: str = "chi_sim+eng",
                 ocr_engine: str = "auto",
                 dpi: int = 300,
                 cache_dir: str = ".wallstreet/qyyjt_cache"):
        self.ocr_lang = ocr_lang
        self.ocr_engine = ocr_engine
        self.dpi = dpi
        self.cache_dir = Path(cache_dir)
        self.cache_dir.mkdir(parents=True, exist_ok=True)

        self._lock = threading.RLock()
        self._stats = {"processed": 0, "errors": 0, "ocr_calls": 0}
        self._check_deps()

    def _check_deps(self):
        """检测可用依赖"""
        self._has_pypdf2 = self._try_import("PyPDF2")
        self._has_pdfplumber = self._try_import("pdfplumber")
        self._has_pillow = self._try_import("PIL")
        self._has_pdf2image = self._try_import("pdf2image")
        self._has_openpyxl = self._try_import("openpyxl")
        self._has_tesseract = self._check_tesseract()
        self._has_paddleocr = self._try_import("paddleocr")

    @staticmethod
    def _try_import(name: str) -> bool:
        try:
            __import__(name)
            return True
        except ImportError:
            return False

    @staticmethod
    def _check_tesseract() -> bool:
        try:
            result = subprocess.run(
                ["tesseract", "--version"],
                capture_output=True, text=True, timeout=5
            )
            return result.returncode == 0
        except Exception:
            return False

    # ═══════════════════════════════════════════════════════
    # 主入口
    # ═══════════════════════════════════════════════════════

    def process_url(self, url: str, headers: Optional[Dict] = None) -> ProcessedDocument:
        """从URL下载并处理文档"""
        import time
        start = time.time()

        content_type = self._detect_content_type(url)

        # 缓存检查
        cache_key = self._cache_key(url)
        cached = self._cache_get(cache_key)
        if cached:
            return cached

        try:
            resp = requests.get(url, headers=headers or {}, timeout=60)
            resp.raise_for_status()
            content = resp.content

            # HTTP头可能提供更准确的类型
            http_ct = resp.headers.get("content-type", "")
            if "pdf" in http_ct:
                content_type = ContentType.PDF
            elif "image" in http_ct:
                content_type = ContentType.IMAGE

            doc = self.process_bytes(content, url.split("/")[-1] or "document", content_type)

        except Exception as e:
            doc = ProcessedDocument(
                source_url=url,
                content_type=content_type,
                error=str(e),
            )
            self._stats["errors"] += 1

        doc.processing_time = round(time.time() - start, 2)
        self._stats["processed"] += 1

        # 缓存
        if not doc.error:
            self._cache_set(cache_key, doc)

        return doc

    def process_bytes(self, data: bytes, filename: str = "document",
                       hint_type: Optional[ContentType] = None) -> ProcessedDocument:
        """处理字节流"""
        content_type = hint_type or self._detect_content_type(filename, data)

        if content_type == ContentType.PDF:
            return self._process_pdf_bytes(data, filename)
        elif content_type == ContentType.IMAGE:
            return self._process_image_bytes(data, filename)
        elif content_type == ContentType.HTML:
            return self._process_html(data, filename)
        elif content_type == ContentType.EXCEL:
            return self._process_excel(data, filename)
        else:
            # 尝试当文本处理
            try:
                text = data.decode("utf-8", errors="replace")
                return ProcessedDocument(
                    source_url=filename,
                    content_type=ContentType.TEXT,
                    total_pages=1,
                    full_text=text,
                    pages=[ExtractedPage(page_number=1, text=text)],
                )
            except Exception:
                return ProcessedDocument(
                    source_url=filename,
                    content_type=ContentType.UNKNOWN,
                    error="Unable to decode content",
                )

    def extract_from_response(self, api_response: Dict[str, Any]) -> List[ProcessedDocument]:
        """
        从API响应中智能提取所有可处理的文档。

        企业预警通API响应可能包含:
          - file_url: 直接PDF下载链接
          - content: base64编码的文档
          - html: HTML格式的表格数据
        """
        docs = []

        # 递归搜索所有可能的文档URL
        urls = self._find_urls_in_dict(api_response)

        for url in urls[:10]:  # 最多处理10个
            try:
                doc = self.process_url(url)
                docs.append(doc)
            except Exception:
                pass

        return docs

    # ═══════════════════════════════════════════════════════
    # 类型检测
    # ═══════════════════════════════════════════════════════

    def _detect_content_type(self, filename_or_url: str,
                              data: Optional[bytes] = None) -> ContentType:
        """检测文件类型"""
        # URL/文件名后缀
        lower = filename_or_url.lower()
        if lower.endswith(".pdf"):
            return ContentType.PDF
        if lower.endswith((".png", ".jpg", ".jpeg", ".tiff", ".tif", ".bmp", ".gif")):
            return ContentType.IMAGE
        if lower.endswith((".xls", ".xlsx", ".csv")):
            return ContentType.EXCEL
        if lower.endswith((".html", ".htm")):
            return ContentType.HTML
        if lower.endswith((".txt", ".md")):
            return ContentType.TEXT

        # 魔数检测
        if data:
            if data[:4] == b"%PDF":
                return ContentType.PDF
            if data[:3] == b"\xff\xd8\xff":
                return ContentType.IMAGE  # JPEG
            if data[:8] == b"\x89PNG\r\n\x1a\n":
                return ContentType.IMAGE  # PNG
            if data[:2] in (b"II", b"MM"):
                return ContentType.IMAGE  # TIFF

        return ContentType.UNKNOWN

    # ═══════════════════════════════════════════════════════
    # PDF 处理
    # ═══════════════════════════════════════════════════════

    def _process_pdf_bytes(self, data: bytes, filename: str) -> ProcessedDocument:
        """处理PDF字节流"""
        doc = ProcessedDocument(
            source_url=filename,
            content_type=ContentType.PDF,
        )
        texts = []

        # PyPDF2
        if self._has_pypdf2:
            try:
                import PyPDF2
                reader = PyPDF2.PdfReader(io.BytesIO(data))
                doc.total_pages = len(reader.pages)

                for i, page in enumerate(reader.pages):
                    text = (page.extract_text() or "").strip()
                    texts.append(text)
                    doc.pages.append(ExtractedPage(
                        page_number=i + 1,
                        text=text,
                        is_scanned=len(text) < 100,
                    ))
            except Exception as e:
                doc.error += f"; PyPDF2: {e}"

        # pdfplumber (表格提取)
        if self._has_pdfplumber and doc.total_pages == 0:
            try:
                import pdfplumber
                with pdfplumber.open(io.BytesIO(data)) as pdf:
                    doc.total_pages = len(pdf.pages)
                    for i, page in enumerate(pdf.pages):
                        if i < len(doc.pages):
                            tables = page.extract_tables()
                            if tables:
                                doc.pages[i].tables = tables
                        text = (page.extract_text() or "").strip()
                        if text and text not in " ".join(texts):
                            texts.append(text)
            except Exception as e:
                doc.error += f"; pdfplumber: {e}"

        doc.full_text = "\n".join(texts)

        # 扫描件OCR
        scanned = [p for p in doc.pages if p.is_scanned]
        if scanned and (self._has_tesseract or self._has_paddleocr):
            with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as f:
                f.write(data)
                tmp_pdf = f.name
            try:
                for page in scanned:
                    ocr_text = self._ocr_pdf_page(tmp_pdf, page.page_number)
                    if ocr_text:
                        page.text = ocr_text
                        page.is_scanned = False
                        texts.append(ocr_text)
            finally:
                Path(tmp_pdf).unlink(missing_ok=True)

        doc.full_text = "\n".join(t for t in texts if t)
        doc.structured_data = self._extract_all_structured(
            doc.full_text, filename
        )

        return doc

    # ═══════════════════════════════════════════════════════
    # 图片处理
    # ═══════════════════════════════════════════════════════

    def _process_image_bytes(self, data: bytes, filename: str) -> ProcessedDocument:
        """处理图片字节流 (OCR)"""
        doc = ProcessedDocument(
            source_url=filename,
            content_type=ContentType.IMAGE,
            total_pages=1,
        )

        # 保存临时文件 (OCR工具需要文件路径)
        with tempfile.NamedTemporaryFile(
            suffix=Path(filename).suffix or ".png", delete=False
        ) as f:
            f.write(data)
            tmp_img = f.name

        try:
            text = self._ocr_image(tmp_img)
            if text:
                self._stats["ocr_calls"] += 1

            doc.pages.append(ExtractedPage(
                page_number=1,
                text=text,
                is_scanned=True,
                confidence=0.7,
            ))
            doc.full_text = text
            doc.structured_data = self._extract_all_structured(text, filename)
        except Exception as e:
            doc.error = str(e)
            self._stats["errors"] += 1
        finally:
            Path(tmp_img).unlink(missing_ok=True)

        return doc

    # ═══════════════════════════════════════════════════════
    # HTML / Excel 处理
    # ═══════════════════════════════════════════════════════

    def _process_html(self, data: bytes, filename: str) -> ProcessedDocument:
        """处理HTML (提取文本+表格)"""
        text = data.decode("utf-8", errors="replace")

        # 提取所有表格
        tables = []
        table_pattern = re.compile(r"<table[^>]*>(.*?)</table>", re.DOTALL | re.IGNORECASE)
        for match in table_pattern.finditer(text):
            rows = []
            row_pattern = re.compile(r"<tr[^>]*>(.*?)</tr>", re.DOTALL | re.IGNORECASE)
            for row_match in row_pattern.finditer(match.group(1)):
                cells = re.findall(r"<t[dh][^>]*>(.*?)</t[dh]>", row_match.group(1), re.DOTALL | re.IGNORECASE)
                rows.append([re.sub(r"<[^>]+>", "", c).strip() for c in cells])
            if rows:
                tables.append(rows)

        # 纯文本
        plain = re.sub(r"<[^>]+>", " ", text)
        plain = re.sub(r"\s+", " ", plain).strip()

        doc = ProcessedDocument(
            source_url=filename,
            content_type=ContentType.HTML,
            total_pages=1,
            full_text=plain,
        )
        if tables:
            doc.pages.append(ExtractedPage(
                page_number=1, text=plain[:500], tables=tables
            ))

        return doc

    def _process_excel(self, data: bytes, filename: str) -> ProcessedDocument:
        """处理Excel文件"""
        if not self._has_openpyxl:
            return ProcessedDocument(
                source_url=filename, content_type=ContentType.EXCEL,
                error="openpyxl not installed",
            )

        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
            texts = []
            all_tables = []

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                table = []
                for row in ws.iter_rows(values_only=True):
                    table.append([str(c) if c is not None else "" for c in row])
                if table:
                    all_tables.append(table)
                    texts.append(f"[Sheet: {sheet_name}]\n" + "\n".join(
                        "\t".join(row) for row in table[:50]  # 每sheet最多50行
                    ))

            return ProcessedDocument(
                source_url=filename,
                content_type=ContentType.EXCEL,
                total_pages=len(wb.sheetnames),
                full_text="\n\n".join(texts),
                pages=[ExtractedPage(page_number=1, text="\n\n".join(texts), tables=all_tables)],
            )
        except Exception as e:
            return ProcessedDocument(
                source_url=filename, content_type=ContentType.EXCEL, error=str(e)
            )

    # ═══════════════════════════════════════════════════════
    # OCR
    # ═══════════════════════════════════════════════════════

    def _ocr_pdf_page(self, pdf_path: str, page_num: int) -> str:
        """对PDF单页OCR"""
        if not self._has_pdf2image:
            return ""

        try:
            from pdf2image import convert_from_path
            images = convert_from_path(
                pdf_path, first_page=page_num, last_page=page_num, dpi=self.dpi
            )
            if images:
                with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as f:
                    images[0].save(f.name, "PNG")
                    tmp = f.name
                text = self._ocr_image(tmp)
                Path(tmp).unlink(missing_ok=True)
                return text
        except Exception:
            pass
        return ""

    def _ocr_image(self, image_path: str) -> str:
        """图片OCR — Tesseract优先，PaddleOCR兜底"""
        if self._has_tesseract:
            result = self._ocr_tesseract(image_path)
            if result:
                return result

        if self._has_paddleocr:
            result = self._ocr_paddleocr(image_path)
            if result:
                return result

        return ""

    def _ocr_tesseract(self, image_path: str) -> str:
        try:
            result = subprocess.run(
                ["tesseract", image_path, "stdout", "-l", self.ocr_lang, "--psm", "6"],
                capture_output=True, text=True, timeout=30,
            )
            if result.returncode == 0:
                return result.stdout.strip()
        except Exception:
            pass
        return ""

    def _ocr_paddleocr(self, image_path: str) -> str:
        try:
            from paddleocr import PaddleOCR
            ocr = PaddleOCR(lang="ch", show_log=False)
            result = ocr.ocr(image_path)
            if result and result[0]:
                return "\n".join(line[1][0] for line in result[0])
        except Exception:
            pass
        return ""

    # ═══════════════════════════════════════════════════════
    # 结构化提取
    # ═══════════════════════════════════════════════════════

    def _extract_all_structured(self, text: str, filename: str) -> Dict[str, Any]:
        """从文本中提取所有可识别的结构化信息 (多类型识别)"""
        if not text.strip():
            return {}

        data = {"document_type": "unknown", "multimodal_version": "v1.1.0"}

        # 优先级匹配: 从最具体的文档类型开始
        checks = [
            # 处罚决定书
            (["行政处罚决定书", "处罚决定书", "行政处罚", "违法事实", "处罚机关"],
             "penalty_decision", lambda: vars(self.extract_penalty_decision(text))),
            # 审计报告
            (["审计报告", "审计意见", "关键审计事项", "会计师事务所"],
             "audit_opinion", lambda: vars(self.extract_audit_opinion(text))),
            # 债券募集说明书
            (["募集说明书", "债券名称", "发行人", "发行规模", "主承销商"],
             "bond_prospectus", lambda: vars(self.extract_bond_prospectus(text))),
            # 执行裁定书
            (["执行裁定书", "执行案号", "申请执行人", "被执行人", "执行标的"],
             "execution_ruling", lambda: vars(self.extract_execution_ruling(text))),
            # 失信决定书
            (["失信被执行人", "失信决定书", "失信名单", "限制消费"],
             "dishonesty_notice", lambda: vars(self.extract_dishonesty_notice(text))),
            # 信用评级
            (["信用评级", "主体信用等级", "债项信用等级", "评级展望"],
             "credit_rating", lambda: vars(self.extract_credit_rating(text))),
            # 营业执照
            (["营业执照", "统一社会信用代码", "法定代表人"],
             "business_license", lambda: vars(self.extract_license_data(text))),
            # 裁判文书 (在处罚/执行之后 — 避免误判)
            (["民事判决书", "刑事判决书", "行政判决书", "裁定书"],
             "court_document", lambda: vars(self.extract_court_data(text))),
            # 财务报表
            (["资产负债表", "利润表", "现金流量表", "营业收入", "净利润"],
             "financial_statement", lambda: vars(self.extract_financial_data(text))),
        ]

        for keywords, dtype, extractor in checks:
            if any(kw in text for kw in keywords):
                data["document_type"] = dtype
                data.update(extractor())
                break  # 第一个匹配胜出

        # 通用提取 (所有类型都提取)
        dates = re.findall(r"\d{4}[-/年]\d{1,2}[-/月]\d{1,2}[日号]?", text)
        if dates:
            data["dates_found"] = dates[:5]

        amounts = re.findall(r"(?:¥|￥|CNY|RMB)\s*[\d,]+\.?\d*\s*(?:万|亿|元)?", text)
        if amounts:
            data["amounts_found"] = amounts[:5]

        # 风险关键词扫描
        risk_keywords = [
            "违约", "逾期", "失信", "执行", "破产", "清算", "吊销",
            "处罚", "警告", "整改", "暂停", "限制", "冻结", "查封",
            "重大不利变化", "无法表示意见", "否定意见", "持续经营能力存在重大不确定性",
        ]
        data["risk_signals"] = [kw for kw in risk_keywords if kw in text][:10]

        return data

    def extract_license_data(self, text: str) -> EnterpriseLicense:
        """从文本中提取营业执照信息"""
        lic = EnterpriseLicense(raw_text=text)

        patterns = {
            "uscc": r"统一社会信用代码[：:]\s*([\dA-Z]{18})",
            "company_name": r"(?:名称|公司名称|企业名称)[：:]\s*([^\n]{4,50})",
            "legal_person": r"法定代表人[：:]\s*([^\n]{2,10})",
            "registered_capital": r"注册资本[：:]\s*([^\n]{2,30})",
            "established_date": r"成立日期[：:]\s*(\d{4}[-/年]\d{1,2}[-/月]\d{1,2})",
            "business_scope": r"经营范围[：:]\s*([^\n]{10,300})",
            "address": r"(?:住所|地址|营业场所)[：:]\s*([^\n]{5,100})",
        }

        for field, pattern in patterns.items():
            m = re.search(pattern, text)
            if m:
                setattr(lic, field, m.group(1).strip())

        return lic

    def extract_court_data(self, text: str) -> CourtDocument:
        """从裁判文书中提取关键字段"""
        doc = CourtDocument(raw_text=text)

        patterns = {
            "case_number": r"[（(]\d{4}[）)][\u4e00-\u9fff]+[\u4e00-\u9fff]*\d+号",
            "court": r"(?:审理法院|本院)[：:]?\s*([^\n]{4,20}(?:法院|仲裁委))",
            "plaintiff": r"(?:原告|申请人|公诉机关)[：:]?\s*([^\n,，]{2,30})",
            "defendant": r"(?:被告|被申请人)[：:]?\s*([^\n,，]{2,30})",
            "filing_date": r"(?:立案日期|受理日期)[：:]\s*(\d{4}[-/年]\d{1,2}[-/月]\d{1,2})",
            "judgment_date": r"(?:判决日期|裁定日期)[：:]\s*(\d{4}[-/年]\d{1,2}[-/月]\d{1,2})",
            "judgment_amount": r"(?:涉案金额|标的额|执行标的)[：:]\s*([^\n]{2,30})",
        }

        for field, pattern in patterns.items():
            m = re.search(pattern, text)
            if m:
                setattr(doc, field, m.group(1).strip())

        return doc

    def extract_financial_data(self, text: str) -> FinancialIndicators:
        """从财务报表中提取关键指标"""
        fin = FinancialIndicators(raw_text=text)

        patterns = {
            "total_assets": r"总资产[：:]\s*([\d,]+\.?\d*)",
            "total_liabilities": r"总负债[：:]\s*([\d,]+\.?\d*)",
            "revenue": r"(?:营业收入|主营业务收入)[：:]\s*([\d,]+\.?\d*)",
            "net_profit": r"(?:净利润|归属于.*净利润)[：:]\s*([\d,]+\.?\d*)",
            "debt_ratio": r"资产负债率[：:]\s*([\d.]+%?)",
            "current_ratio": r"流动比率[：:]\s*([\d.]+%?)",
            "roe": r"(?:ROE|净资产收益率)[：:]\s*([\d.]+%?)",
            "operating_cashflow": r"经营活动.*现金流[：:]\s*([\d,]+\.?\d*)",
        }

        for field, pattern in patterns.items():
            m = re.search(pattern, text)
            if m:
                setattr(fin, field, m.group(1).strip())

        # 年份
        year_match = re.search(r"(\d{4})\s*年(?:度|报|财)", text)
        if year_match:
            fin.year = year_match.group(1)

        return fin

    # ═══════════════════════════════════════════════════════
    # 扩展文档类型提取 — 处罚/审计/债券/执行/失信/评级
    # ═══════════════════════════════════════════════════════

    def extract_penalty_decision(self, text: str) -> PenaltyDecision:
        """提取行政处罚决定书"""
        pd = PenaltyDecision(raw_text=text)
        pat = {
            "penalty_number": r"(?:行政处罚决定书|处罚决定书)\s*(?:文号|编号)?[：:]?\s*([^\n]{5,40})",
            "agency": r"(?:处罚机关|作出处罚|决定机关)[：:]?\s*([^\n]{4,30})",
            "target": r"(?:被处罚人|当事人|被处罚单位)[：:]?\s*([^\n,，]{4,50})",
            "violation": r"(?:违法事实|违法行为|违法情形)[：:]?\s*([^\n]{10,200})",
            "penalty_type": r"(?:处罚种类|处罚类型)[：:]?\s*([^\n]{4,50})",
            "penalty_amount": r"(?:罚款|处罚金额|并处)[：:]?\s*([^\n]{2,30})",
            "decision_date": r"(?:处罚日期|决定日期|作出日期)[：:]?\s*(\d{4}[-/年]\d{1,2}[-/月]\d{1,2})",
            "legal_basis": r"(?:依据|法律依据)[：:]?\s*([^\n]{10,100})",
        }
        for f, p in pat.items():
            m = re.search(p, text); 
            if m: setattr(pd, f, m.group(1).strip())
        return pd

    def extract_audit_opinion(self, text: str) -> AuditOpinion:
        """提取审计报告/审计意见"""
        ao = AuditOpinion(raw_text=text)
        pat = {
            "auditor": r"(?:审计机构|会计师事务所|审计单位)[：:]?\s*([^\n]{4,30})",
            "audit_type": r"(?:审计意见类型|意见类型)[：:]?\s*([^\n]{2,20})",
            "audit_year": r"(?:审计年度|报告年度)[：:]?\s*(\d{4})",
            "key_matters": r"(?:关键审计事项|强调事项)[：:]?\s*([^\n]{10,300})",
            "going_concern": r"(?:持续经营|持续经营能力).*?([^\n]{5,100})",
            "report_date": r"(?:审计报告日|报告日期|出具日期)[：:]?\s*(\d{4}[-/年]\d{1,2}[-/月]\d{1,2})",
        }
        for f, p in pat.items():
            m = re.search(p, text); 
            if m: setattr(ao, f, m.group(1).strip())
        # 自动判断意见类型
        if not ao.audit_type:
            if "无法表示意见" in text: ao.audit_type = "无法表示意见"
            elif "否定意见" in text: ao.audit_type = "否定意见"
            elif "保留意见" in text: ao.audit_type = "保留意见"
            elif "无保留意见" in text or "标准无保留" in text: ao.audit_type = "标准无保留意见"
        return ao

    def extract_bond_prospectus(self, text: str) -> BondProspectus:
        """提取债券募集说明书"""
        bp = BondProspectus(raw_text=text)
        pat = {
            "bond_name": r"(?:债券名称|债券简称)[：:]?\s*([^\n]{4,40})",
            "issuer": r"(?:发行人|发行主体)[：:]?\s*([^\n,，]{4,50})",
            "bond_type": r"(?:债券品种|债券类型)[：:]?\s*([^\n]{4,30})",
            "issue_amount": r"(?:发行规模|发行金额|发行总额)[：:]?\s*([^\n]{2,30})",
            "maturity": r"(?:债券期限|期限)[：:]?\s*([^\n]{2,20})",
            "coupon_rate": r"(?:票面利率|利率)[：:]?\s*([^\n]{2,15})",
            "credit_rating": r"(?:信用评级|债项评级|主体评级)[：:]?\s*([^\n]{2,20})",
            "lead_underwriter": r"(?:主承销商|牵头主承)[：:]?\s*([^\n]{4,30})",
            "use_of_proceeds": r"(?:募集资金用途|资金用途)[：:]?\s*([^\n]{10,200})",
            "guarantor": r"(?:担保人|担保方|增信方)[：:]?\s*([^\n]{4,30})",
        }
        for f, p in pat.items():
            m = re.search(p, text); 
            if m: setattr(bp, f, m.group(1).strip())
        return bp

    def extract_execution_ruling(self, text: str) -> ExecutionRuling:
        """提取执行裁定书"""
        er = ExecutionRuling(raw_text=text)
        pat = {
            "case_number": r"(?:执行案号|执行裁定书)[：:]?\s*([（(]\d{4}[）)][^\n]*?\d+号)",
            "court": r"(?:执行法院|本院)[：:]?\s*([^\n]{4,20}(?:法院|执行局))",
            "applicant": r"(?:申请执行人|申请人)[：:]?\s*([^\n,，]{4,30})",
            "respondent": r"(?:被执行人|被申请人)[：:]?\s*([^\n,，]{4,30})",
            "execution_amount": r"(?:执行标的|执行金额|申请执行标的)[：:]?\s*([^\n]{2,30})",
            "filing_date": r"(?:申请执行日期|立案日期)[：:]?\s*(\d{4}[-/年]\d{1,2}[-/月]\d{1,2})",
            "ruling_date": r"(?:裁定日期|作出日期)[：:]?\s*(\d{4}[-/年]\d{1,2}[-/月]\d{1,2})",
            "execution_status": r"(?:执行情况|执行结果)[：:]?\s*([^\n]{5,100})",
            "property_found": r"(?:财产发现|查控财产|财产线索)[：:]?\s*([^\n]{5,100})",
        }
        for f, p in pat.items():
            m = re.search(p, text); 
            if m: setattr(er, f, m.group(1).strip())
        return er

    def extract_dishonesty_notice(self, text: str) -> DishonestyNotice:
        """提取失信被执行人决定书"""
        dn = DishonestyNotice(raw_text=text)
        pat = {
            "case_number": r"(?:案号|执行案号)[：:]?\s*([（(]\d{4}[）)][^\n]*?\d+号)",
            "court": r"(?:执行法院|发布法院)[：:]?\s*([^\n]{4,20}(?:法院|人民法院))",
            "dishonest_person": r"(?:失信被执行人|被执行人)[：:]?\s*([^\n,，]{4,30})",
            "id_number": r"(?:身份证号|统一社会信用代码|组织机构代码)[：:]?\s*([^\n]{10,30})",
            "violation": r"(?:失信情形|失信行为|具体情形)[：:]?\s*([^\n]{5,200})",
            "publish_date": r"(?:发布日期|公布日期|列入日期)[：:]?\s*(\d{4}[-/年]\d{1,2}[-/月]\d{1,2})",
            "legal_basis": r"(?:法律依据|依据)[：:]?\s*([^\n]{5,100})",
        }
        for f, p in pat.items():
            m = re.search(p, text); 
            if m: setattr(dn, f, m.group(1).strip())
        return dn

    def extract_credit_rating(self, text: str) -> CreditRating:
        """提取信用评级报告"""
        cr = CreditRating(raw_text=text)
        pat = {
            "rated_entity": r"(?:评级对象|受评主体|发行人)[：:]?\s*([^\n,，]{4,50})",
            "rating_agency": r"(?:评级机构|评级公司)[：:]?\s*([^\n]{4,30})",
            "subject_rating": r"(?:主体信用等级|主体评级)[：:]?\s*([^\n]{2,10})",
            "bond_rating": r"(?:债项信用等级|债项评级)[：:]?\s*([^\n]{2,10})",
            "outlook": r"(?:评级展望|展望)[：:]?\s*([^\n]{2,10})",
            "rating_date": r"(?:评级日期|报告日期)[：:]?\s*(\d{4}[-/年]\d{1,2}[-/月]\d{1,2})",
            "key_factors": r"(?:评级关注|主要关注|风险因素)[：:]?\s*([^\n]{10,200})",
            "upgrade_downgrade": r"(?:评级调整|调整方向|调整幅度)[：:]?\s*([^\n]{5,50})",
        }
        for f, p in pat.items():
            m = re.search(p, text); 
            if m: setattr(cr, f, m.group(1).strip())
        return cr

    # ═══════════════════════════════════════════════════════
    # 工具方法
    # ═══════════════════════════════════════════════════════

    def _find_urls_in_dict(self, obj: Any, depth: int = 0) -> List[str]:
        """递归查找字典中所有URL"""
        if depth > 5:
            return []

        urls = []
        if isinstance(obj, dict):
            for k, v in obj.items():
                if isinstance(v, str) and (
                    v.startswith("http") and
                    any(ext in v.lower() for ext in [".pdf", ".png", ".jpg", ".xls", ".xlsx", ".doc"])
                ):
                    urls.append(v)
                elif isinstance(v, (dict, list)):
                    urls.extend(self._find_urls_in_dict(v, depth + 1))
        elif isinstance(obj, list):
            for item in obj:
                urls.extend(self._find_urls_in_dict(item, depth + 1))

        return urls

    def _cache_key(self, url: str) -> str:
        import hashlib
        return hashlib.md5(url.encode()).hexdigest()[:16]

    def _cache_get(self, key: str) -> Optional[ProcessedDocument]:
        cache_file = self.cache_dir / f"{key}.json"
        if cache_file.is_file():
            try:
                data = json.loads(cache_file.read_text(encoding="utf-8"))
                return ProcessedDocument(**data)
            except Exception:
                pass
        return None

    def _cache_set(self, key: str, doc: ProcessedDocument):
        cache_file = self.cache_dir / f"{key}.json"
        try:
            cache_file.write_text(
                json.dumps(doc, default=str, ensure_ascii=False, indent=2),
                encoding="utf-8",
            )
        except Exception:
            pass

    def get_stats(self) -> dict:
        return dict(self._stats)

    @property
    def available_features(self) -> List[str]:
        """列出当前环境可用的处理能力"""
        features = ["text_extraction"]
        if self._has_pypdf2:
            features.append("pdf_text")
        if self._has_pdfplumber:
            features.append("pdf_table")
        if self._has_tesseract:
            features.append("ocr_tesseract")
        if self._has_paddleocr:
            features.append("ocr_paddleocr")
        if self._has_openpyxl:
            features.append("excel")
        if self._has_pdf2image:
            features.append("pdf_render")
        return features


# ═══════════════════════════════════════════════════════════════
# QYJT Adapter 集成
# ═══════════════════════════════════════════════════════════════

class QYJTWithMultimodal:
    """
    企业预警通 + 多模态一体查询

    用法:
        qm = QYJTWithMultimodal()
        result = qm.full_query("特斯拉")
        # result["documents"] 包含所有自动处理的 PDF/图片
        # result["licenses"] 包含提取的营业执照信息
    """

    def __init__(self):
        from adapters.qyyjt_adapter import QYJTAdapter
        self.adapter = QYJTAdapter()
        self.processor = QYJTMultimodalProcessor()

    def full_query(self, company: str) -> Dict[str, Any]:
        """全功能查询: 数据 + 多模态处理"""
        result = {
            "company": company,
            "timestamp": datetime.now().isoformat(),
            "search_results": {},
            "documents": [],
            "licenses": [],
            "court_docs": [],
            "penalties": [],
            "audit_opinions": [],
            "bond_prospectuses": [],
            "execution_rulings": [],
            "dishonesty_notices": [],
            "credit_ratings": [],
            "financials": [],
        }

        # 搜索
        try:
            result["search_results"] = self.adapter.search_company(company)
        except Exception as e:
            result["search_error"] = str(e)

        # 从搜索结果中提取文件URL并处理
        all_urls = self.processor._find_urls_in_dict(result["search_results"])

        # 分类路由 — 将文档分派给对应的Agent角色
        doc_route = {
            "business_license": ("licenses", "张铁柱·企业尽调"),
            "court_document": ("court_docs", "赵刚·风险评估"),
            "penalty_decision": ("penalties", "赵刚·风险评估"),
            "execution_ruling": ("execution_rulings", "赵刚·风险评估"),
            "dishonesty_notice": ("dishonesty_notices", "赵刚·风险评估"),
            "financial_statement": ("financials", "李明远·财务分析"),
            "audit_opinion": ("audit_opinions", "李明远·财务分析"),
            "bond_prospectus": ("bond_prospectuses", "王思远·行业研究"),
            "credit_rating": ("credit_ratings", "王思远·行业研究"),
        }

        for url in all_urls[:10]:
            try:
                doc = self.processor.process_url(url)
                st = doc.structured_data
                dt = st.get("document_type", "unknown")

                record = {
                    "url": url,
                    "type": doc.content_type.value,
                    "document_type": dt,
                    "text_preview": doc.full_text[:500],
                    "structured": st,
                    "risk_signals": st.get("risk_signals", []),
                    "assignee": doc_route.get(dt, ("documents", "郑慎之·交叉验证"))[1],
                }

                result["documents"].append(record)

                # 自动路由到正确的角色分类
                target_list, _ = doc_route.get(dt, ("documents", ""))
                if target_list in result:
                    result[target_list].append(st)
                else:
                    result["documents"][-1]["structured"] = st

            except Exception as e:
                result["documents"].append({"url": url, "error": str(e)})

        # 汇总: 生成各Agent的任务摘要
        result["agent_assignments"] = {
            "赵刚·风险评估": {
                "court_cases": len(result["court_docs"]),
                "penalties": len(result["penalties"]),
                "executions": len(result["execution_rulings"]),
                "dishonesty": len(result["dishonesty_notices"]),
                "total": len(result["court_docs"]) + len(result["penalties"]) +
                         len(result["execution_rulings"]) + len(result["dishonesty_notices"]),
            },
            "李明远·财务分析": {
                "financials": len(result["financials"]),
                "audit_opinions": len(result["audit_opinions"]),
                "total": len(result["financials"]) + len(result["audit_opinions"]),
            },
            "张铁柱·企业尽调": {
                "licenses": len(result["licenses"]),
            },
            "王思远·行业研究": {
                "bond_prospectuses": len(result["bond_prospectuses"]),
                "credit_ratings": len(result["credit_ratings"]),
                "total": len(result["bond_prospectuses"]) + len(result["credit_ratings"]),
            },
            "郑慎之·交叉验证": {
                "unrouted": sum(1 for d in result["documents"] if d.get("document_type") == "unknown"),
                "note": "未识别的文档类型需要人工判断后分配",
            },
        }

        return result
